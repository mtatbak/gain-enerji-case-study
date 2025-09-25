import xlsxwriter
from datetime import datetime, timedelta
import calendar
import pandas as pd
from openpyxl import load_workbook

def create_excel_template(filename, santral1='Santral_1', santral2='Santral_2'):
    workbook = xlsxwriter.Workbook(filename)
    
    header_format = workbook.add_format({
        'bold': True,
        'bg_color': '#D7E4BC',
        'border': 1,
        'align': 'center',
        'valign': 'vcenter',
        'text_wrap': True
    })
    
    date_format = workbook.add_format({
        'num_format': 'dd/mm/yyyy',
        'border': 1,
        'align': 'center'
    })
    
    number_format = workbook.add_format({
        'border': 1,
        'align': 'center'
    })
    
    formula_format = workbook.add_format({
        'border': 1,
        'align': 'center',
        'num_format': '#,##0.00'
    })
    
    headers = [
        'Tarih', 'Ay', 'Saat','PTF', 'SMF','Pozitif Den. Fiyatı', 'Negatif Den. Fiyatı',
        'Gün Öncesi Üretim Tahmini (KGÜP)', 'Gerçekleşen Üretim',
        'Dengesizlik Miktarı', 'GÖP Geliri (TL)',
        'Dengesizlik Tutarı  (TL)', 'Toplam (Net) Gelir (TL)',
        'Dengesizlik Maliyeti (TL)', 'Birim Dengesizlik Maliyeti (TL/MWh)']
    
    #Santral1
    
    ws1 = workbook.add_worksheet(santral1)
    
    for col, header in enumerate(headers):
        ws1.write(0, col, header, header_format)
    
    ws1.set_column('A:A', 12)  # Tarih 
    ws1.set_column('B:B', 6)   # Ay 
    ws1.set_column('C:C', 6)   # Saat 
    ws1.set_column(3, len(headers)-1, 15)  
    
    
    current_row = 2
    total_hours = 0
    
    for month in range(1, 13):  
        last_day = calendar.monthrange(2024, month)[1] 
        for day in range(1, last_day + 1):  
            for hour in range(24):  
                date = datetime(2024, month, day)
                
                ws1.write(current_row, 0, date, date_format)      # A : Tarih
                ws1.write(current_row, 1, month, number_format)   # B : Ay
                ws1.write(current_row, 2, hour, number_format)    # C : Saat

                for col in range(3, len(headers)):
                    ws1.write(current_row, col, '', number_format)
                
                current_row += 1
                total_hours += 1
    
    
    #Santral2
    
    ws2 = workbook.add_worksheet(santral2)
    
    for col, header in enumerate(headers):
        ws2.write(0, col, header, header_format)
    
    ws2.set_column('A:A', 12)
    ws2.set_column('B:B', 6)
    ws2.set_column('C:C', 6)
    ws2.set_column(3, len(headers)-1, 15)
    
    
    current_row = 2
    
    for month in range(1, 13):
        last_day = calendar.monthrange(2024, month)[1]
        for day in range(1, last_day + 1):
            for hour in range(24):
                date = datetime(2024, month, day)
                
                ws2.write(current_row, 0, date, date_format)
                ws2.write(current_row, 1, month, number_format)
                ws2.write(current_row, 2, hour, number_format)
                
                for col in range(3, len(headers)):
                    ws2.write(current_row, col, '', number_format)
                
                current_row += 1
    
    
    #Karşılaştırma
    
    ws3 = workbook.add_worksheet('Karşılaştırma')

    comparison_headers = [
        'Ay',
        'Gerçekleşen Üretim (MWh)',
        'Dengesizlik Miktarı (MWh)',
        'GÖP Geliri (TL)',
        'Dengesizlik Tutarı (TL)',
        'Toplam Gelir (TL)',
        'Birim Gelir (TL/MWh)',
        'Dengesizlik Maliyeti (TL)',
        'Birim Deng. Mal. (TL/MWh)'
    ]
    
    red_bold_format = workbook.add_format({
        'bold': True,
        'font_color': 'red'
    })

    header_format_green = workbook.add_format({
        'bold': True,
        'bg_color': '#D7E4BC',
        'border': 1,
        'align': 'center',
        'valign': 'vcenter',
        'text_wrap': True
    })

    bordered_center_format = workbook.add_format({
        'border': 1,
        'align': 'center',
        'valign': 'vcenter'
    })

    # Santral1 
    
    ws3.write('B2', f'{santral1}', red_bold_format)

    for col, header in enumerate(comparison_headers):
        ws3.write(3, col + 1, header, header_format_green)  

    for month in range(1, 13):  
        row = 3 + month  
        ws3.write(row, 1, month, bordered_center_format)  
        
        # Excel satır numarası (Python 0-based, Excel 1-based)
        excel_row = row + 1
        
        safe_santral1 = f"'{santral1}'" if ' ' in santral1 else santral1
        
        # Gerçekleşen Üretim (MWh) - I sütunu (index 8)
        formula = f"=SUMIF({safe_santral1}!B:B,{month},{safe_santral1}!I:I)"
        ws3.write_formula(row, 2, formula, formula_format)
        
        # Dengesizlik Miktarı (MWh) - J sütunu (index 9)
        formula = f"=SUMIF({safe_santral1}!B:B,{month},{safe_santral1}!J:J)"
        ws3.write_formula(row, 3, formula, formula_format)
        
        # GÖP Geliri (TL) - K sütunu (index 10)
        formula = f"=SUMIF({safe_santral1}!B:B,{month},{safe_santral1}!K:K)"
        ws3.write_formula(row, 4, formula, formula_format)
        
        # Dengesizlik Tutarı (TL) - L sütunu (index 11)
        formula = f"=SUMIF({safe_santral1}!B:B,{month},{safe_santral1}!L:L)"
        ws3.write_formula(row, 5, formula, formula_format)
        
        # Toplam Gelir (TL) - M sütunu (index 12)
        formula = f"=SUMIF({safe_santral1}!B:B,{month},{safe_santral1}!M:M)"
        ws3.write_formula(row, 6, formula, formula_format)
        
        # Birim Gelir (TL/MWh) - Toplam Gelir / Gerçekleşen Üretim
        formula = f"=IF(C{excel_row}=0,0,G{excel_row}/C{excel_row})"
        ws3.write_formula(row, 7, formula, formula_format)
        
        # Dengesizlik Maliyeti (TL) - N sütunu (index 13)
        formula = f"=SUMIF({safe_santral1}!B:B,{month},{safe_santral1}!N:N)"
        ws3.write_formula(row, 8, formula, formula_format)
        
        # Birim Deng. Mal. (TL/MWh) - O sütunu (index 14)
        formula = f"=SUMIF({safe_santral1}!B:B,{month},{safe_santral1}!O:O)"
        ws3.write_formula(row, 9, formula, formula_format)

    ws3.write(16, 1, 'Toplam', header_format_green)
    
    for col in range(2, 10):  
        if col == 7:  # Birim Gelir için ağırlıklı ortalama
            formula = f"=IF(SUM(C5:C16)=0,0,SUM(G5:G16)/SUM(C5:C16))"
        elif col == 9:  # Birim Maliyet için ağırlıklı ortalama
            formula = f"=IF(SUM(C5:C16)=0,0,SUM(I5:I16)/SUM(C5:C16))"
        else: 
            col_letter = chr(ord('C') + col - 2)  # C, D, E, F, G, H, I, J
            formula = f"=SUM({col_letter}5:{col_letter}16)"
        
        ws3.write_formula(16, col, formula, formula_format)

    #Santral2
    ws3.write('B19', f'{santral2}', red_bold_format)

    for col, header in enumerate(comparison_headers):
        ws3.write(20, col + 1, header, header_format_green)  

    for month in range(1, 13):  
        row = 20 + month  
        ws3.write(row, 1, month, bordered_center_format)  
        
        excel_row = row + 1
        
        safe_santral2 = f"'{santral2}'" if ' ' in santral2 else santral2
        
        # Gerçekleşen Üretim (MWh)
        formula = f"=SUMIF({safe_santral2}!B:B,{month},{safe_santral2}!I:I)"
        ws3.write_formula(row, 2, formula, formula_format)
        
        # Dengesizlik Miktarı (MWh)
        formula = f"=SUMIF({safe_santral2}!B:B,{month},{safe_santral2}!J:J)"
        ws3.write_formula(row, 3, formula, formula_format)
        
        # GÖP Geliri (TL)
        formula = f"=SUMIF({safe_santral2}!B:B,{month},{safe_santral2}!K:K)"
        ws3.write_formula(row, 4, formula, formula_format)
        
        # Dengesizlik Tutarı (TL)
        formula = f"=SUMIF({safe_santral2}!B:B,{month},{safe_santral2}!L:L)"
        ws3.write_formula(row, 5, formula, formula_format)
        
        # Toplam Gelir (TL)
        formula = f"=SUMIF({safe_santral2}!B:B,{month},{safe_santral2}!M:M)"
        ws3.write_formula(row, 6, formula, formula_format)
        
        # Birim Gelir (TL/MWh)
        formula = f"=IF(C{excel_row}=0,0,G{excel_row}/C{excel_row})"
        ws3.write_formula(row, 7, formula, formula_format)
        
        # Dengesizlik Maliyeti (TL)
        formula = f"=SUMIF({safe_santral2}!B:B,{month},{safe_santral2}!N:N)"
        ws3.write_formula(row, 8, formula, formula_format)
        
        # Birim Deng. Mal. (TL/MWh)
        formula = f"=SUMIF({safe_santral2}!B:B,{month},{safe_santral2}!O:O)"
        ws3.write_formula(row, 9, formula, formula_format)

    ws3.write(33, 1, 'Toplam', header_format_green)
    
    for col in range(2, 10):  
        if col == 7:  # Birim Gelir için ağırlıklı ortalama
            formula = f"=IF(SUM(C22:C33)=0,0,SUM(G22:G33)/SUM(C22:C33))"
        elif col == 9:  # Birim Maliyet için ağırlıklı ortalama
            formula = f"=IF(SUM(C22:C33)=0,0,SUM(I22:I33)/SUM(C22:C33))"
        else:  
            col_letter = chr(ord('C') + col - 2)
            formula = f"=SUM({col_letter}22:{col_letter}33)"
        
        ws3.write_formula(33, col, formula, formula_format)

    ws3.set_column('B:J', 20)
    
    workbook.close()

def load_data_excel(df, excel_path, sheet_name, mapping):
    
    wb = load_workbook(excel_path)
    ws = wb[sheet_name]

    # Excel'in ilk satırındaki başlıkları oku
    headers = [cell.value for cell in ws[1]]  
    
    # Mapping'deki her sütun için işlem yap
    for df_column, excel_header in mapping.items():
        if excel_header not in headers:
            continue
        
        col_index = headers.index(excel_header) + 1 
     
        start_row = 3
        
        for i, value in enumerate(df[df_column]):
            excel_row = start_row + i  # 3, 4, 5, 6...
            ws.cell(row=excel_row, column=col_index, value=value)
    
    wb.save(excel_path)
    wb.close()

